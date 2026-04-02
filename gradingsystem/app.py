import os
import smtplib
import re
import pandas as pd
import sqlite3
import psycopg2
from psycopg2.extras import RealDictCursor
from email.message import EmailMessage
from flask import Flask, render_template, request, redirect, url_for, jsonify, session, send_file
import time
import io

app = Flask(__name__)
app.secret_key = 'prakalp-secure-iot-key-2026'

# 🌐 UNIVERSAL EMAIL VALIDATOR
EMAIL_REGEX = re.compile(r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$')

# ==============================
# 🔐 CONFIG
# ==============================
SENDER_EMAIL = 'prakalp-hackathon@rcee.ac.in'
SENDER_PASSWORD = 'zlbrpnprgnynuzso'
ADMIN_USERNAME = 'admin'
ADMIN_PASSWORD = 'iotsprint@#1234'

# Create necessary directories
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# 🚀 Vercel fix: Use /tmp for writable files in serverless environment
if os.environ.get('VERCEL'):
    TMP_DIR = '/tmp'
else:
    TMP_DIR = os.path.join(BASE_DIR, 'data', 'tmp')
    os.makedirs(TMP_DIR, exist_ok=True)

REGISTRY_PATH = os.path.join(TMP_DIR, 'registry.csv')

def get_db_connection():
    url = os.environ.get('POSTGRES_URL') or os.environ.get('STORAGE_URL')
    if url:
        try:
            return psycopg2.connect(url, sslmode='require')
        except Exception as e:
            print(f"❌ Postgres error: {e}")
            return None
    else:
        # Fallback to local SQLite for development
        try:
            conn = sqlite3.connect(os.path.join(BASE_DIR, 'data', 'gradingsystem.db'))
            conn.row_factory = sqlite3.Row
            return conn
        except Exception as e:
            print(f"❌ SQLite error: {e}")
            return None

# ==============================
# 📊 SCHEMA & METADATA
# ==============================
ROUNDS = ['round1', 'round2', 'round3p1', 'round3p2', 'round4']

GRADE_POINTS = {
    'A+': 100,
    'A': 90,
    'B+': 80,
    'B': 70,
    'C': 60,
    '': 0
}

def initialize_db(path=None, wipe=False):
    c = get_db_connection()
    if not c: return
    try:
        cur = c.cursor()
        schema = """CREATE TABLE IF NOT EXISTS grades_teams (teamid TEXT PRIMARY KEY, projectid TEXT, teamname TEXT, projecttitle TEXT, email TEXT,
        round1 TEXT DEFAULT '', round2 TEXT DEFAULT '', round3p1 TEXT DEFAULT '', round3p2 TEXT DEFAULT '', round4 TEXT DEFAULT '');"""
        cur.execute(schema)
        
        if wipe or path:
            print("Wiping teams table...")
            cur.execute("DELETE FROM grades_teams;")
            c.commit() 
            
        if path:
            print(f"Initializing from: {path}")
            df = pd.read_csv(path, sep=None, engine='python', encoding_errors='ignore') if str(path).lower().endswith('.csv') else pd.read_excel(path)
            df.columns = [str(col).strip() for col in df.columns]
            
            for r in df.to_dict(orient='records'):
                try:
                    def f(ks, exclude=None):
                        ks = [k.lower() for k in ks]
                        for rk in r.keys():
                            rk_clean = str(rk).lower().strip()
                            if exclude and any(ex.lower() in rk_clean for ex in exclude): continue
                            if any(k in rk_clean for k in ks): return str(r[rk]).strip()
                        return ""
                    
                    email = f(['Email', 'Mail', 'Id'], exclude=['@']) # exclude columns that aren't the email itself
                    if "@" not in email: # try harder finding it in any field
                        for val in r.values():
                            if isinstance(val, str) and "@" in val:
                                email = val.strip()
                                break
                    
                    # 🚀 FALLBACK: If still no email, use a placeholder to allow import
                    if not email or "@" not in email:
                        roll = f(['Roll', 'Number', 'Reg'])
                        name_slug = f(['Name', 'Student']).lower().replace(" ", "")
                        email = f"{roll or name_slug}@prakalp-hackathon.com"

                    tid = f(['Batch', 'TeamID', 'ProjectID', 'S.No']) or "0"
                    pid = f(['Batch', 'ProjectID', 'PID']) or "PR-IOT"
                    name = f(['Name', 'Student', 'Team']) or "Anonymous Team"
                    title = f(['Title', 'Problem', 'Statement']) or "Untitled Project"
                        
                    if isinstance(c, sqlite3.Connection):
                        cur.execute("INSERT OR REPLACE INTO grades_teams (teamid, projectid, teamname, projecttitle, email) VALUES (?,?,?,?,?)", (tid, pid, name, title, email))
                    else:
                        cur.execute("INSERT INTO grades_teams (teamid, projectid, teamname, projecttitle, email) VALUES (%s,%s,%s,%s,%s) ON CONFLICT (teamid) DO UPDATE SET projectid=EXCLUDED.projectid, teamname=EXCLUDED.teamname, projecttitle=EXCLUDED.projecttitle, email=EXCLUDED.email;", (tid, pid, name, title, email))
                except Exception as row_err:
                    print(f"⚠️ Row error: {row_err}")
                    continue
            
            c.commit() # Final commit for inserts
            print("Import finished.")
        else:
            c.commit() # Normal schema creation commit
    finally: c.close()

def get_teams():
    c = get_db_connection()
    if not c: return []
    try:
        if isinstance(c, sqlite3.Connection):
            raw = c.execute("SELECT * FROM grades_teams;").fetchall()
            res = [dict(r) for r in raw]
        else:
            cur = c.cursor(cursor_factory=RealDictCursor)
            cur.execute("SELECT * FROM grades_teams;")
            res = [dict(r) for r in cur.fetchall()]
        
        for t in res:
            # Consistent mapping with safe fallbacks
            t['TeamID'] = t.get('teamid') or str(t.get('TeamID') or '')
            t['ProjectID'] = t.get('projectid') or str(t.get('ProjectID') or 'PR-IOT')
            t['TeamName'] = t.get('teamname') or str(t.get('TeamName') or 'Anonymous Team')
            t['ProjectTitle'] = t.get('projecttitle') or str(t.get('ProjectTitle') or 'Untitled Project')
            t['Email'] = t.get('email') or str(t.get('Email') or '')
            
            # Map letters
            r1 = t.get('round1') or ''
            r2 = t.get('round2') or ''
            r3p1 = t.get('round3p1') or ''
            r3p2 = t.get('round3p2') or ''
            r4 = t.get('round4') or ''
            
            t['R1_Grade'], t['R2_Grade'], t['R3P1_Grade'], t['R3P2_Grade'], t['R4_Grade'] = r1, r2, r3p1, r3p2, r4
            
            # Retrieve points
            p1 = GRADE_POINTS.get(r1.upper(), 0)
            p2 = GRADE_POINTS.get(r2.upper(), 0)
            p3p1 = GRADE_POINTS.get(r3p1.upper(), 0)
            p3p2 = GRADE_POINTS.get(r3p2.upper(), 0)
            p4 = GRADE_POINTS.get(r4.upper(), 0)

            # 🏆 Final Combined Score
            t['Weighted_Total'] = (p1*0.15) + (p2*0.15) + (p3p1*0.2) + (p3p2*0.2) + (p4*0.3)
            
            # Keep raw total around just in case
            t['Raw_Total'] = p1 + p2 + p3p1 + p3p2 + p4

        return sorted(res, key=lambda x: x.get('Weighted_Total', 0), reverse=True)
    except Exception as e:
        print(f"❌ Error in get_teams: {e}")
        initialize_db()
        return []
    finally: c.close()

# ==============================
# 📧 ENGINE & ROUTES
# ==============================
def send_email(to, sub, body, server=None):
    try:
        msg = EmailMessage()
        msg.set_content(body)
        msg['Subject'] = sub
        msg['From'] = f"PRAKALP 2026<{SENDER_EMAIL}>"
        msg['To'] = to.strip()

        if server:
            server.send_message(msg)
        else:
            with smtplib.SMTP_SSL('smtp.gmail.com', 465) as s:
                s.login(SENDER_EMAIL, SENDER_PASSWORD)
                s.send_message(msg)
        return True, None
    except Exception as e:
        return False, str(e)

@app.route('/')
def index():
    if not session.get('admin_logged_in'): return redirect(url_for('login'))
    return render_template('index.html', teams=get_teams())

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        if request.form.get('username')==ADMIN_USERNAME and request.form.get('password')==ADMIN_PASSWORD:
            session['admin_logged_in'] = True
            return redirect(url_for('admin'))
        else:
            return render_template('login.html', error="Invalid username or password. Access Denied.")
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.pop('admin_logged_in', None)
    return redirect(url_for('index'))

@app.route('/admin')
def admin():
    if not session.get('admin_logged_in'): return redirect(url_for('login'))
    round_info = {
        'round1': ('Round 1', 'R1_Grade'),
        'round2': ('Round 2', 'R2_Grade'),
        'round3p1': ('Round 3 P1', 'R3P1_Grade'),
        'round3p2': ('Round 3 P2', 'R3P2_Grade'),
        'round4': ('Round 4', 'R4_Grade')
    }
    
    # Sort strictly by Team ID ascending for consistent judge input
    t_list = get_teams()
    t_list.sort(key=lambda x: (len(str(x.get('TeamID', ''))), str(x.get('TeamID', ''))))
    
    return render_template('admin.html', teams=t_list, rounds=round_info, grades=['A+', 'A', 'B+', 'B', 'C'])

@app.route('/update_scores', methods=['POST'])
def update_scores():
    if not session.get('admin_logged_in'): return redirect(url_for('login'))
    fd, c = request.form.to_dict(), get_db_connection()
    if not c: return redirect(url_for('admin'))
    try:
        cur = c.cursor()
        for sno in set([k.split('_')[-1] for k in fd.keys() if '_' in k]):
            up, pa = [], []
            for rnd in ROUNDS:
                field_key = f"{rnd}_{sno}"
                if field_key in fd:
                    up.append(f"{rnd} = ?" if isinstance(c, sqlite3.Connection) else f"{rnd} = %s")
                    pa.append(fd[field_key].strip().upper())
            if up:
                pa.append(sno)
                query = f"UPDATE grades_teams SET {', '.join(up)} WHERE teamid = ?" if isinstance(c, sqlite3.Connection) else f"UPDATE grades_teams SET {', '.join(up)} WHERE teamid = %s"
                cur.execute(query, pa)
        c.commit()
    finally: c.close()
    return redirect(url_for('admin') + '?saved=1')


@app.route('/upload_dispatch', methods=['POST'])
def upload_dispatch():
    if not session.get('admin_logged_in'): return redirect(url_for('login'))
    file = request.files.get('registry_file')
    if not file: return redirect(url_for('admin'))
    file.save(REGISTRY_PATH)
    
    try:
        df = pd.read_csv(REGISTRY_PATH, sep=None, engine='python', encoding_errors='ignore') if REGISTRY_PATH.endswith('.csv') else pd.read_excel(REGISTRY_PATH)
        # 🚀 CLEAN HEADERS: Ensure no whitespace issues
        df.columns = [str(c).strip() for c in df.columns]
        # 🚀 NO EMAILS: User requested to just initialize team only
        initialize_db(REGISTRY_PATH, wipe=True)
        return redirect(url_for('admin') + '?registered=1&tab=setup')
    except Exception as e:
        print(f"❌ Dispatch error: {e}")
        return redirect(url_for('admin') + f'?error=dispatch_failed&reason={str(e)}&tab=setup')

@app.route('/finalize_registry', methods=['POST'])
def finalize_registry():
    if not session.get('admin_logged_in'): return redirect(url_for('login'))
    if os.path.exists(REGISTRY_PATH):
        initialize_db(REGISTRY_PATH)
    import time
    return redirect(url_for('admin') + f'?registered=1&tab=setup&t={int(time.time())}')

@app.route('/download_results')
def download_results():
    if not session.get('admin_logged_in'): return redirect(url_for('login'))
    teams = get_teams()
    if not teams: return "No data found."
    
    output = io.BytesIO()
    try:
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            fdf = pd.DataFrame(teams)
            if fdf.empty: 
                 return "Registry is empty."
                 
            fdf.insert(0, 'Rank', range(1, len(fdf) + 1))
            
            def assign_prize(rank):
                if rank == 1: return "🥇 1st Prize"
                if rank == 2: return "🥈 2nd Prize"
                if rank <= 5: return "🥉 3rd Prize"
                return ""
            
            fdf.insert(1, 'Prize Category', fdf['Rank'].apply(assign_prize))
            
            final_mapping = {
                'Rank': 'Rank',
                'Prize Category': 'Award',
                'TeamID': 'Team ID',
                'ProjectID': 'Project ID',
                'TeamName': 'Team Name',
                'ProjectTitle': 'Project Title',
                'R1_Grade': 'Round 1',
                'R2_Grade': 'Round 2',
                'R3P1_Grade': 'Round 3 P1',
                'R3P2_Grade': 'Round 3 P2',
                'R4_Grade': 'Round 4',
                'Weighted_Total': 'Weighted Score (%)'
            }
            
            fexport = fdf.reindex(columns=list(final_mapping.keys()), fill_value='')
            fexport.rename(columns=final_mapping, inplace=True)
            fexport.to_excel(writer, index=False, sheet_name='Final Standings')
            
            ws = writer.sheets['Final Standings']
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True, size=11)
            center_align = Alignment(horizontal="center", vertical="center")
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_align
                cell.border = thin_border

            gold_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
            silver_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            bronze_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                rank_val = row[0].value
                fill = None
                if rank_val == 1: fill = gold_fill
                elif rank_val == 2: fill = silver_fill
                elif rank_val <= 5: fill = bronze_fill
                
                for cell in row:
                    if fill: cell.fill = fill
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical="center", horizontal="left" if cell.column in [3, 5, 6] else "center")

            for col in ws.columns:
                max_len = max([len(str(cell.value)) for cell in col])
                ws.column_dimensions[col[0].column_letter].width = max_len + 4

    except Exception as e:
        return f"Excel Generation Error: {str(e)}"

    output.seek(0)
    return send_file(output, as_attachment=True, download_name='PRAKALP_GradingSystem_Results.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    output.seek(0)
    return send_file(output, as_attachment=True, download_name='PRAKALP_2026_IoT_Hackathon_Final_Results.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/reset_db', methods=['POST'])
def reset_db():
    if not session.get('admin_logged_in'): return redirect(url_for('login'))
    initialize_db(wipe=True)
    return redirect(url_for('admin') + '?reset=1&tab=setup')

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000)