import os
import smtplib
import re
import pandas as pd
import sqlite3
import psycopg2
from psycopg2.extras import RealDictCursor
from email.message import EmailMessage
from flask import Flask, render_template, request, redirect, url_for, jsonify, session, send_file
import time
import io

app = Flask(__name__)
app.secret_key = 'prakalp-secure-iot-key-2026'

# 🌐 UNIVERSAL EMAIL VALIDATOR
EMAIL_REGEX = re.compile(r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$')

# ==============================
# 🔐 CONFIG
# ==============================
SENDER_EMAIL = 'prakalp-hackathon@rcee.ac.in'
SENDER_PASSWORD = 'zlbrpnprgnynuzso'
ADMIN_USERNAME = 'admin'
ADMIN_PASSWORD = 'iotsprint@#1234'

# Create necessary directories
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# 🚀 Vercel fix: Use /tmp for writable files in serverless environment
if os.environ.get('VERCEL'):
    TMP_DIR = '/tmp'
else:
    TMP_DIR = os.path.join(BASE_DIR, 'data', 'tmp')
    os.makedirs(TMP_DIR, exist_ok=True)

REGISTRY_PATH = os.path.join(TMP_DIR, 'registry.csv')

def get_db_connection():
    url = os.environ.get('POSTGRES_URL') or os.environ.get('STORAGE_URL')
    if url:
        try:
            return psycopg2.connect(url, sslmode='require')
        except Exception as e:
            print(f"❌ Postgres error: {e}")
            return None
    else:
        # 📂 Fallback to local SQLite for development
        try:
            conn = sqlite3.connect(os.path.join(BASE_DIR, 'data', 'hackathon.db'))
            conn.row_factory = sqlite3.Row
            return conn
        except Exception as e:
            print(f"❌ SQLite error: {e}")
            return None

# ==============================
# 📊 SCHEMA & METADATA
# ==============================
SCORING_FIELDS = {
    'R1': ['r1_innovation', 'r1_problemrelevance', 'r1_techfeasibility', 'r1_claritypresentation'],
    'R2': ['r2_feasibilityvalidation', 'r2_systemdesignlogic', 'r2_demoquality', 'r2_technicalunderstanding'],
    'R3P1': ['r3p1_initialimplementation', 'r3p1_approachmethodology', 'r3p1_progresslevel', 'r3p1_teamcoordination'],
    'R3P2': ['r3p2_improvementphase1', 'r3p2_innovationmodification', 'r3p2_problemsolvingapproach', 'r3p2_stabilityfunctionality'],
    'R4': ['r4_innovation', 'r4_workingprototype', 'r4_realtimeimpact', 'r4_presentationskills', 'r4_qahandling']
}

SCORING_META = {
    'R1': {'title': 'Round 1 – Idea Submission', 'desc': 'PPT pitch (3–5 min)', 'weight': '15%', 'criteria': [{'field': 'r1_innovation', 'label': 'Innovation', 'max': 15}, {'field': 'r1_problemrelevance', 'label': 'Problem Relevance', 'max': 10}, {'field': 'r1_techfeasibility', 'label': 'Technical Feasibility', 'max': 15}, {'field': 'r1_claritypresentation', 'label': 'Clarity of Presentation', 'max': 10}]},
    'R2': {'title': 'Round 2 – Concept Validation', 'desc': 'Simulation/Architecture', 'weight': '15%', 'criteria': [{'field': 'r2_feasibilityvalidation', 'label': 'Feasibility Validation', 'max': 20}, {'field': 'r2_systemdesignlogic', 'label': 'System Design & Logic', 'max': 15}, {'field': 'r2_demoquality', 'label': 'Demonstration Quality', 'max': 10}, {'field': 'r2_technicalunderstanding', 'label': 'Technical Understanding', 'max': 5}]},
    'R3P1': {'title': 'Round 3 – Phase 1', 'desc': 'Early Implementation', 'weight': '20%', 'criteria': [{'field': 'r3p1_initialimplementation', 'label': 'Initial Implementation', 'max': 20}, {'field': 'r3p1_approachmethodology', 'label': 'Approach & Methodology', 'max': 10}, {'field': 'r3p1_progresslevel', 'label': 'Progress Level', 'max': 10}, {'field': 'r3p1_teamcoordination', 'label': 'Team Coordination', 'max': 10}]},
    'R3P2': {'title': 'Round 3 – Phase 2', 'desc': 'Modification Phase', 'weight': '20%', 'criteria': [{'field': 'r3p2_improvementphase1', 'label': 'Improvement from Phase 1', 'max': 15}, {'field': 'r3p2_innovationmodification', 'label': 'Innovation in Modification', 'max': 15}, {'field': 'r3p2_problemsolvingapproach', 'label': 'Problem-Solving Approach', 'max': 10}, {'field': 'r3p2_stabilityfunctionality', 'label': 'Stability & Functionality', 'max': 10}]},
    'R4': {'title': 'Round 4 – Final Pitch', 'desc': 'Live Demo & Q&A', 'weight': '30%', 'criteria': [{'field': 'r4_innovation', 'label': 'Innovation', 'max': 10}, {'field': 'r4_workingprototype', 'label': 'Working Prototype', 'max': 15}, {'field': 'r4_realtimeimpact', 'label': 'Real-Time Impact', 'max': 10}, {'field': 'r4_presentationskills', 'label': 'Presentation Skills', 'max': 5}, {'field': 'r4_qahandling', 'label': 'Q&A Handling', 'max': 10}]}
}

def initialize_db(path=None, wipe=False):
    c = get_db_connection()
    if not c: return
    try:
        cur = c.cursor()
        schema = """CREATE TABLE IF NOT EXISTS teams (teamid TEXT PRIMARY KEY, projectid TEXT, teamname TEXT, projecttitle TEXT, email TEXT,
        r1_innovation NUMERIC DEFAULT 0, r1_problemrelevance NUMERIC DEFAULT 0, r1_techfeasibility NUMERIC DEFAULT 0, r1_claritypresentation NUMERIC DEFAULT 0,
        r2_feasibilityvalidation NUMERIC DEFAULT 0, r2_systemdesignlogic NUMERIC DEFAULT 0, r2_demoquality NUMERIC DEFAULT 0, r2_technicalunderstanding NUMERIC DEFAULT 0,
        r3p1_initialimplementation NUMERIC DEFAULT 0, r3p1_approachmethodology NUMERIC DEFAULT 0, r3p1_progresslevel NUMERIC DEFAULT 0, r3p1_teamcoordination NUMERIC DEFAULT 0,
        r3p2_improvementphase1 NUMERIC DEFAULT 0, r3p2_innovationmodification NUMERIC DEFAULT 0, r3p2_problemsolvingapproach NUMERIC DEFAULT 0, r3p2_stabilityfunctionality NUMERIC DEFAULT 0,
        r4_innovation NUMERIC DEFAULT 0, r4_workingprototype NUMERIC DEFAULT 0, r4_realtimeimpact NUMERIC DEFAULT 0, r4_presentationskills NUMERIC DEFAULT 0, r4_qahandling NUMERIC DEFAULT 0);"""
        cur.execute(schema)
        
        if wipe or path:
            print("Wiping teams table...")
            cur.execute("DELETE FROM teams;")
            c.commit() 
            
        if path:
            print(f"Initializing from: {path}")
            df = pd.read_csv(path, sep=None, engine='python', encoding_errors='ignore') if str(path).lower().endswith('.csv') else pd.read_excel(path)
            df.columns = [str(col).strip() for col in df.columns]
            
            for r in df.to_dict(orient='records'):
                try:
                    def f(ks, exclude=None):
                        ks = [k.lower() for k in ks]
                        for rk in r.keys():
                            rk_clean = str(rk).lower().strip()
                            if exclude and any(ex.lower() in rk_clean for ex in exclude): continue
                            if any(k in rk_clean for k in ks): return str(r[rk]).strip()
                        return ""
                    
                    email = f(['Email', 'Mail', 'Id'], exclude=['@']) # exclude columns that aren't the email itself
                    if "@" not in email: # try harder finding it in any field
                        for val in r.values():
                            if isinstance(val, str) and "@" in val:
                                email = val.strip()
                                break
                    
                    # 🚀 FALLBACK: If still no email, use a placeholder to allow import
                    if not email or "@" not in email:
                        roll = f(['Roll', 'Number', 'Reg'])
                        name_slug = f(['Name', 'Student']).lower().replace(" ", "")
                        email = f"{roll or name_slug}@prakalp-hackathon.com"

                    tid = f(['Batch', 'TeamID', 'ProjectID', 'S.No']) or "0"
                    pid = f(['Batch', 'ProjectID', 'PID']) or "PR-IOT"
                    name = f(['Name', 'Student', 'Team']) or "Anonymous Team"
                    title = f(['Title', 'Problem', 'Statement']) or "Untitled Project"
                        
                    if isinstance(c, sqlite3.Connection):
                        cur.execute("INSERT OR REPLACE INTO teams (teamid, projectid, teamname, projecttitle, email) VALUES (?,?,?,?,?)", (tid, pid, name, title, email))
                    else:
                        cur.execute("INSERT INTO teams (teamid, projectid, teamname, projecttitle, email) VALUES (%s,%s,%s,%s,%s) ON CONFLICT (teamid) DO UPDATE SET projectid=EXCLUDED.projectid, teamname=EXCLUDED.teamname, projecttitle=EXCLUDED.projecttitle, email=EXCLUDED.email;", (tid, pid, name, title, email))
                except Exception as row_err:
                    print(f"⚠️ Row error: {row_err}")
                    continue
            
            c.commit() # Final commit for inserts
            print("Import finished.")
        else:
            c.commit() # Normal schema creation commit
    finally: c.close()

def get_teams():
    c = get_db_connection()
    if not c: return []
    try:
        if isinstance(c, sqlite3.Connection):
            raw = c.execute("SELECT * FROM teams;").fetchall()
            res = [dict(r) for r in raw]
        else:
            cur = c.cursor(cursor_factory=RealDictCursor)
            cur.execute("SELECT * FROM teams;")
            res = [dict(r) for r in cur.fetchall()]
        
        for t in res:
            # Consistent mapping with safe fallbacks
            t['TeamID'] = t.get('teamid') or str(t.get('TeamID') or '')
            t['ProjectID'] = t.get('projectid') or str(t.get('ProjectID') or 'PR-IOT')
            t['TeamName'] = t.get('teamname') or str(t.get('TeamName') or 'Anonymous Team')
            t['ProjectTitle'] = t.get('projecttitle') or str(t.get('ProjectTitle') or 'Untitled Project')
            t['Email'] = t.get('email') or str(t.get('Email') or '')
            for k in list(t.keys()):
                if k.startswith('r'): t[k] = float(t.get(k) or 0)
            
            def s(fs): return sum(float(t.get(f) or 0) for f in fs)
            t['R1_Total'], t['R2_Total'], t['R3P1_Total'], t['R3P2_Total'], t['R4_Total'] = s(SCORING_FIELDS['R1']), s(SCORING_FIELDS['R2']), s(SCORING_FIELDS['R3P1']), s(SCORING_FIELDS['R3P2']), s(SCORING_FIELDS['R4'])
            # 🏆 Final Combined Score
            t['Weighted_Total'] = (t['R1_Total']*0.3) + (t['R2_Total']*0.3) + (t['R3P1_Total']*0.4) + (t['R3P2_Total']*0.4) + (t['R4_Total']*0.6)
            t['Raw_Total'] = t['R1_Total'] + t['R2_Total'] + t['R3P1_Total'] + t['R3P2_Total'] + t['R4_Total']

        return sorted(res, key=lambda x: x.get('Weighted_Total', 0), reverse=True)
    except Exception as e:
        print(f"❌ Error in get_teams: {e}")
        initialize_db()
        return []
    finally: c.close()

# ==============================
# 📧 ENGINE & ROUTES
# ==============================
def send_email(to, sub, body, server=None):
    try:
        msg = EmailMessage()
        msg.set_content(body)
        msg['Subject'] = sub
        msg['From'] = f"PRAKALP 2026<{SENDER_EMAIL}>"
        msg['To'] = to.strip()

        if server:
            server.send_message(msg)
        else:
            with smtplib.SMTP_SSL('smtp.gmail.com', 465) as s:
                s.login(SENDER_EMAIL, SENDER_PASSWORD)
                s.send_message(msg)
        return True, None
    except Exception as e:
        return False, str(e)

@app.route('/')
def index(): return render_template('index.html', teams=get_teams())

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        if request.form.get('username')==ADMIN_USERNAME and request.form.get('password')==ADMIN_PASSWORD:
            session['admin_logged_in'] = True
            return redirect(url_for('admin'))
    return render_template('login.html')

@app.route('/admin')
def admin():
    if not session.get('admin_logged_in'): return redirect(url_for('login'))
    return render_template('admin.html', teams=get_teams(), fields=SCORING_FIELDS, meta=SCORING_META)

@app.route('/update_scores', methods=['POST'])
def update_scores():
    if not session.get('admin_logged_in'): return redirect(url_for('login'))
    fd, c = request.form.to_dict(), get_db_connection()
    if not c: return redirect(url_for('admin'))
    try:
        cur = c.cursor()
        for sno in set([k.split('_')[-1] for k in fd.keys() if '_' in k]):
            up, pa = [], []
            for fs in SCORING_FIELDS.values():
                for f in fs:
                    if f"{f}_{sno}" in fd:
                        up.append(f"{f} = %s" if not isinstance(c, sqlite3.Connection) else f"{f} = ?")
                        pa.append(float(fd[f"{f}_{sno}"] or 0))
            if up:
                pa.append(sno)
                query = f"UPDATE teams SET {', '.join(up)} WHERE teamid = %s" if not isinstance(c, sqlite3.Connection) else f"UPDATE teams SET {', '.join(up)} WHERE teamid = ?"
                cur.execute(query, pa)
        c.commit()
    finally: c.close()
    return redirect(url_for('admin') + '?saved=1')

@app.route('/upload_dispatch', methods=['POST'])
def upload_dispatch():
    if not session.get('admin_logged_in'): return redirect(url_for('login'))
    file = request.files.get('registry_file')
    if not file: return redirect(url_for('admin'))
    file.save(REGISTRY_PATH)
    
    try:
        df = pd.read_csv(REGISTRY_PATH, sep=None, engine='python', encoding_errors='ignore') if REGISTRY_PATH.endswith('.csv') else pd.read_excel(REGISTRY_PATH)
        # 🚀 CLEAN HEADERS: Ensure no whitespace issues
        df.columns = [str(c).strip() for c in df.columns]
        # 🚀 NO EMAILS: User requested to just initialize team only
        initialize_db(REGISTRY_PATH, wipe=True)
        return redirect(url_for('admin') + f'?registered=1&tab=setup')
    except Exception as e:
        print(f"❌ Dispatch error: {e}")
        return redirect(url_for('admin') + f'?error=dispatch_failed&reason={str(e)}&tab=setup')

@app.route('/finalize_registry', methods=['POST'])
def finalize_registry():
    if not session.get('admin_logged_in'): return redirect(url_for('login'))
    if os.path.exists(REGISTRY_PATH):
        initialize_db(REGISTRY_PATH)
    import time
    return redirect(url_for('admin') + f'?registered=1&tab=setup&t={int(time.time())}')

# ... (rest of the file handles Excel generation and rankings)

from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

@app.route('/download_results')
def download_results():
    if not session.get('admin_logged_in'): return redirect(url_for('login'))
    teams = get_teams()
    if not teams: return "No data found."
    
    output = io.BytesIO()
    try:
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # 🧪 1-5: Detailed Round Sheets
            for rkey, rmeta in SCORING_META.items():
                if rkey == 'setup': continue
                fields = [c['field'] for c in rmeta['criteria']]
                cols = ['teamid', 'teamname'] + fields + [f'{rkey.lower()}_total']
                
                # Use reindex to ensure all columns exist even if NULL in DB
                rdf = pd.DataFrame(teams).reindex(columns=cols, fill_value=0)
                
                # Friendly Header Names
                header_map = {'teamid': 'Team ID', 'teamname': 'Team Name', f'{rkey.lower()}_total': 'Total Score'}
                for c in rmeta['criteria']: header_map[c['field']] = c['label']
                rdf.rename(columns=header_map, inplace=True)
                
                # Excel Sheet name limit is 31 chars
                s_name = rmeta['title'][:31]
                rdf.to_excel(writer, index=False, sheet_name=s_name)
                
                # 🔥 Auto-Width adjustment
                ws = writer.sheets[s_name]
                for col in ws.columns:
                    max_len = max([len(str(cell.value)) for cell in col])
                    ws.column_dimensions[col[0].column_letter].width = max_len + 4

            # 🏆 6: FINAL PRINT-READY STANDINGS
            fdf = pd.DataFrame(teams)
            if fdf.empty: 
                 return "Registry is empty."
                 
            fdf.insert(0, 'Rank', range(1, len(fdf) + 1))
            
            def assign_prize(rank):
                if rank == 1: return "🥇 1st Prize"
                if rank == 2: return "🥈 2nd Prize"
                if rank <= 5: return "🥉 3rd Prize"
                return ""
            
            fdf.insert(1, 'Prize Category', fdf['Rank'].apply(assign_prize))
            
            # Map database keys to friendly export names
            final_mapping = {
                'Rank': 'Rank',
                'Prize Category': 'Award',
                'teamid': 'Team ID',
                'projectid': 'Project ID',
                'teamname': 'Team Name',
                'projecttitle': 'Project Title',
                'r1_total': 'R1 (15%)',
                'r2_total': 'R2 (15%)',
                'r3p1_total': 'R3P1 (20%)',
                'r3p2_total': 'R3P2 (20%)',
                'r4_total': 'R4 (30%)',
                'weighted_total': 'Weighted Total (%)'
            }
            
            fexport = fdf.reindex(columns=list(final_mapping.keys()), fill_value=0)
            fexport.rename(columns=final_mapping, inplace=True)
            fexport.to_excel(writer, index=False, sheet_name='Final Standings')
            
            # 🎨 Management-Standard Formatting
            ws = writer.sheets['Final Standings']
            header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True, size=11)
            center_align = Alignment(horizontal="center", vertical="center")
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_align
                cell.border = thin_border

            # Row highlights
            gold_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
            silver_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            bronze_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                rank_val = row[0].value
                fill = None
                if rank_val == 1: fill = gold_fill
                elif rank_val == 2: fill = silver_fill
                elif rank_val <= 5: fill = bronze_fill
                
                for cell in row:
                    if fill: cell.fill = fill
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical="center", horizontal="left" if cell.column in [3, 5, 6] else "center")

            widths = {'A':6, 'B':15, 'C':15, 'D':15, 'E':25, 'F':35}
            for col, width in widths.items():
                ws.column_dimensions[col].width = width

    except Exception as e:
        return f"Excel Generation Error: {str(e)}"

    output.seek(0)
    return send_file(output, as_attachment=True, download_name='PRAKALP_2026_IoT_Hackathon_Final_Results.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    output.seek(0)
    return send_file(output, as_attachment=True, download_name='PRAKALP_2026_IoT_Hackathon_Final_Results.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/reset_db', methods=['POST'])
def reset_db():
    if not session.get('admin_logged_in'): return redirect(url_for('login'))
    initialize_db(wipe=True)
    return redirect(url_for('admin') + '?reset=1&tab=setup')

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000)